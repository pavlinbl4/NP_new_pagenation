"""
Из скачанных парсером снимков нового проспекта за месяц генерю
файл отчета с превью
"""
from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import (
    Border, Side,
    Alignment, Font
)
from openpyxl.utils import get_column_letter

from gui_tools.select_month import select_month
from parsing.image_resize import image_resize
from xlsx_tools.set_column_dimensions import set_column_dimensions

month_name = select_month()
current_year = datetime.now().year
icloud_folder = Path().home() / 'Library/Mobile Documents/com~apple~CloudDocs/'

(icloud_folder / f"Documents/NewProspect/{current_year}_{month_name}").mkdir(parents=True, exist_ok=True)
way_to_files = (icloud_folder / f"Documents/NewProspect/{current_year}_{month_name}")

workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Sheet with image"  # задаю название вкладки

# in list widths of all columns
worksheet = set_column_dimensions(worksheet, [15, 38, 21, 5])

thin_border = Border(left=Side(border_style="thin"),
                     right=Side(border_style="thin"),
                     top=Side(border_style="thin"),
                     bottom=Side(border_style="thin"))

row = 0

for row, image_path in enumerate(way_to_files.glob("*.JPG"), 1):
    worksheet.row_dimensions[row].height = 86  # задаю высоту столбца

    logger.info(image_path)

    for column in range(1, 5):
        # cells view
        worksheet[f'{get_column_letter(column)}{row}'].border = thin_border
        worksheet[f'{get_column_letter(column)}{row}'].font = Font(size=10, bold=True)
        worksheet[f'{get_column_letter(column)}{row}'].alignment = Alignment(horizontal='center',
                                                                             vertical='center',
                                                                             wrap_text=True)
        worksheet[f'{get_column_letter(3)}{row}'].alignment = Alignment(horizontal='center',
                                                                        vertical='center',
                                                                        )
        # add information to cells

        worksheet[f'{get_column_letter(1)}{row}'].value = image_path.name.split("__")[0]
        worksheet[f'{get_column_letter(2)}{row}'].value = image_path.name.split("__")[1][:-4]
        img = image_resize(image_path)
        worksheet.add_image(img, f"{get_column_letter(3)}{row}")
        worksheet[f'{get_column_letter(4)}{row}'].value = 500

for column in range(1, 5):
    worksheet[f'{get_column_letter(column)}{row + 3}'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet[f'{get_column_letter(column)}{row + 3}'].font = Font(size=32, bold=True, color="FF0000")
    worksheet[f'{get_column_letter(column)}{row + 3}'].border = thin_border

    worksheet.row_dimensions[row + 3].height = 60
    worksheet[f'{get_column_letter(2)}{row + 3}'].value = "ИТОГО"
    worksheet[f'{get_column_letter(4)}{row + 3}'].value = f"=SUM(D$1:D${row})"

workbook.save(f"{way_to_files}/report_{current_year}_{month_name}.xlsx")
workbook.close()
